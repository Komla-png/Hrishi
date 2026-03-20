"""Excel workbook to database sync utilities.

This module reads analytics data from an Excel workbook and upserts it into
`centers`, `monthly_data`, `coaches`, and `coach_salaries`.
"""

from __future__ import annotations

import logging
import os
import shutil
import sqlite3
import tempfile
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from urllib.parse import unquote, urlparse

from dotenv import load_dotenv
from openpyxl import load_workbook

try:
    import pg8000
except Exception:  # pragma: no cover
    pg8000 = None

CALENDAR_MONTHS = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

LOGGER = logging.getLogger("excel_sync")

# Load .env values if present so scripts work without manual export.
load_dotenv()


@dataclass
class MonthlyDataRow:
    center_name: str
    month: str
    year: int
    revenue: float
    target: float


@dataclass
class CoachSalaryRow:
    center_name: str
    coach_name: str
    month: str
    year: int
    salary: float
    end_month: Optional[str]
    end_year: Optional[int]


@dataclass
class SyncReport:
    monthly_rows_seen: int = 0
    monthly_rows_synced: int = 0
    coach_rows_seen: int = 0
    coach_rows_synced: int = 0
    skipped_rows: int = 0
    warnings: Optional[List[str]] = None

    def as_dict(self) -> dict:
        return {
            "monthly_rows_seen": self.monthly_rows_seen,
            "monthly_rows_synced": self.monthly_rows_synced,
            "coach_rows_seen": self.coach_rows_seen,
            "coach_rows_synced": self.coach_rows_synced,
            "skipped_rows": self.skipped_rows,
            "warnings": self.warnings or [],
        }


class WorksheetMissingError(Exception):
    """Raised when a requested worksheet does not exist in the workbook."""


def _probe_workbook_readable(workbook_path: str) -> None:
    """Open and close workbook once to verify read access."""
    workbook = load_workbook(filename=workbook_path, data_only=True, read_only=True)
    workbook.close()


def _prepare_workbook_for_sync(workbook_path: str) -> Tuple[str, Optional[str]]:
    """Return a readable workbook path and optional temp copy path.

    On Windows, Excel can hold restrictive locks on the live workbook.
    This method first attempts direct access, then falls back to a temporary
    copy to avoid lock contention during sync.
    """
    try:
        _probe_workbook_readable(workbook_path)
        return workbook_path, None
    except PermissionError:
        pass

    try:
        suffix = os.path.splitext(workbook_path)[1] or ".xlsx"
        fd, temp_path = tempfile.mkstemp(prefix="excel_sync_", suffix=suffix)
        os.close(fd)
        shutil.copy2(workbook_path, temp_path)
        _probe_workbook_readable(temp_path)
        return temp_path, temp_path
    except PermissionError as exc:
        raise RuntimeError(
            f"Excel sync file is locked or inaccessible: {workbook_path}. "
            "Close the workbook in Excel and retry."
        ) from exc
    except Exception:
        # Surface original path in errors for easier troubleshooting.
        raise


def _normalize_key(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_")


def _normalize_month(value: object) -> Optional[str]:
    if value is None:
        return None
    month = str(value).strip().title()
    if not month:
        return None
    if month in CALENDAR_MONTHS:
        return month

    short_map = {m[:3].lower(): m for m in CALENDAR_MONTHS}
    return short_map.get(month[:3].lower())


def _safe_float(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    raw = str(value).strip().replace(",", "")
    if raw == "":
        return default
    return float(raw)


def _safe_int(value: object, default: Optional[int] = None) -> Optional[int]:
    if value is None:
        return default
    raw = str(value).strip()
    if raw == "":
        return default
    return int(float(raw))


def read_worksheet_as_records(workbook_path: str, worksheet_title: str) -> List[Dict[str, object]]:
    """Read an Excel worksheet as normalized list of dictionaries."""
    workbook = load_workbook(filename=workbook_path, data_only=True, read_only=True)
    try:
        if worksheet_title not in workbook.sheetnames:
            raise WorksheetMissingError(
                f"Worksheet '{worksheet_title}' not found in {workbook_path}. "
                f"Available sheets: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[worksheet_title]
        rows_iter = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers = [_normalize_key(h) for h in header_row]
        records: List[Dict[str, object]] = []

        for row in rows_iter:
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            item: Dict[str, object] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None
            records.append(item)

        return records
    finally:
        workbook.close()


def parse_monthly_data_rows(records: List[Dict[str, object]]) -> Tuple[List[MonthlyDataRow], List[str]]:
    """Convert raw MonthlyData worksheet rows to validated MonthlyDataRow list."""
    required = {"center_name", "month", "year", "revenue", "target"}
    available = set(records[0].keys()) if records else set()
    missing = required - available
    if records and missing:
        raise ValueError(f"MonthlyData sheet missing required columns: {sorted(missing)}")

    rows: List[MonthlyDataRow] = []
    warnings: List[str] = []

    for idx, item in enumerate(records):
        row_number = idx + 2  # Header is row 1.
        try:
            center_name = str(item.get("center_name", "")).strip()
            if not center_name:
                warnings.append(f"MonthlyData row {row_number}: center_name is empty; skipped")
                continue

            month = _normalize_month(item.get("month"))
            if not month:
                warnings.append(f"MonthlyData row {row_number}: invalid month '{item.get('month')}'; skipped")
                continue

            year = _safe_int(item.get("year"))
            if not year or year < 2000 or year > 2100:
                warnings.append(f"MonthlyData row {row_number}: invalid year '{item.get('year')}'; skipped")
                continue

            revenue = _safe_float(item.get("revenue"), default=0.0)
            target = _safe_float(item.get("target"), default=0.0)

            rows.append(
                MonthlyDataRow(
                    center_name=center_name,
                    month=month,
                    year=year,
                    revenue=revenue,
                    target=target,
                )
            )
        except Exception as exc:
            warnings.append(f"MonthlyData row {row_number}: parse error ({exc}); skipped")

    return rows, warnings


def parse_coach_salary_rows(records: List[Dict[str, object]]) -> Tuple[List[CoachSalaryRow], List[str]]:
    """Convert raw CoachSalaries worksheet rows to validated CoachSalaryRow list."""
    # Recover from a common Excel issue where the first header cell is blank,
    # which gets normalized as the key "none" instead of "center_name".
    if records and "center_name" not in records[0] and "none" in records[0]:
        for item in records:
            if "center_name" not in item and "none" in item:
                item["center_name"] = item.get("none")

    required = {"center_name", "coach_name", "month", "year", "salary"}
    available = set(records[0].keys()) if records else set()
    missing = required - available
    if records and missing:
        raise ValueError(f"CoachSalaries sheet missing required columns: {sorted(missing)}")

    rows: List[CoachSalaryRow] = []
    warnings: List[str] = []

    for idx, item in enumerate(records):
        row_number = idx + 2
        try:
            center_name = str(item.get("center_name", "")).strip()
            coach_name = str(item.get("coach_name", "")).strip()
            if not center_name or not coach_name:
                warnings.append(f"CoachSalaries row {row_number}: center_name/coach_name missing; skipped")
                continue

            month = _normalize_month(item.get("month"))
            if not month:
                warnings.append(f"CoachSalaries row {row_number}: invalid month '{item.get('month')}'; skipped")
                continue

            year = _safe_int(item.get("year"))
            if not year or year < 2000 or year > 2100:
                warnings.append(f"CoachSalaries row {row_number}: invalid year '{item.get('year')}'; skipped")
                continue

            salary = _safe_float(item.get("salary"), default=0.0)
            end_month = _normalize_month(item.get("end_month")) if "end_month" in item else None
            end_year = _safe_int(item.get("end_year")) if "end_year" in item else None

            rows.append(
                CoachSalaryRow(
                    center_name=center_name,
                    coach_name=coach_name,
                    month=month,
                    year=year,
                    salary=salary,
                    end_month=end_month,
                    end_year=end_year,
                )
            )
        except Exception as exc:
            warnings.append(f"CoachSalaries row {row_number}: parse error ({exc}); skipped")

    return rows, warnings


class DatabaseClient:
    """Small DB helper that supports SQLite and PostgreSQL upserts."""

    def __init__(self, conn, backend: str):
        self.conn = conn
        self.cur = conn.cursor()
        self.backend = backend

    @classmethod
    def from_environment(cls) -> "DatabaseClient":
        database_url = os.environ.get("DATABASE_URL", "").strip()

        if database_url.startswith(("postgres://", "postgresql://")):
            if pg8000 is None:
                raise RuntimeError("pg8000 is required for PostgreSQL. Install pg8000.")
            parsed = urlparse(database_url)
            conn = pg8000.connect(
                user=unquote(parsed.username or ""),
                password=unquote(parsed.password or ""),
                host=parsed.hostname or "localhost",
                port=parsed.port or 5432,
                database=(parsed.path or "").lstrip("/") or None,
                ssl_context=True,
            )
            return cls(conn=conn, backend="postgres")

        sqlite_path = os.environ.get("DATABASE_PATH", "instance/academy.db")
        os.makedirs(os.path.dirname(sqlite_path) or ".", exist_ok=True)
        conn = sqlite3.connect(sqlite_path, timeout=30)
        conn.execute("PRAGMA journal_mode=WAL")
        return cls(conn=conn, backend="sqlite")

    def close(self) -> None:
        self.cur.close()
        self.conn.close()

    def _fetch_center_id(self, center_name: str) -> int:
        if self.backend == "sqlite":
            self.cur.execute(
                "INSERT INTO centers(name) VALUES (?) ON CONFLICT(name) DO NOTHING",
                (center_name,),
            )
            self.cur.execute("SELECT id FROM centers WHERE name = ?", (center_name,))
        else:
            self.cur.execute(
                "INSERT INTO centers(name) VALUES (%s) ON CONFLICT(name) DO NOTHING",
                (center_name,),
            )
            self.cur.execute("SELECT id FROM centers WHERE name = %s", (center_name,))

        center_id = self.cur.fetchone()[0]
        return int(center_id)

    def _fetch_coach_id(
        self,
        center_id: int,
        coach_name: str,
        end_month: Optional[str],
        end_year: Optional[int],
    ) -> int:
        if self.backend == "sqlite":
            self.cur.execute(
                """
                INSERT INTO coaches(center_id, name, end_month, end_year)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(center_id, name)
                DO UPDATE SET end_month=excluded.end_month, end_year=excluded.end_year
                """,
                (center_id, coach_name, end_month, end_year),
            )
            self.cur.execute(
                "SELECT id FROM coaches WHERE center_id = ? AND name = ?",
                (center_id, coach_name),
            )
        else:
            self.cur.execute(
                """
                INSERT INTO coaches(center_id, name, end_month, end_year)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT(center_id, name)
                DO UPDATE SET end_month=excluded.end_month, end_year=excluded.end_year
                """,
                (center_id, coach_name, end_month, end_year),
            )
            self.cur.execute(
                "SELECT id FROM coaches WHERE center_id = %s AND name = %s",
                (center_id, coach_name),
            )

        coach_id = self.cur.fetchone()[0]
        return int(coach_id)

    def upsert_monthly_rows(self, rows: List[MonthlyDataRow]) -> int:
        synced = 0
        for row in rows:
            center_id = self._fetch_center_id(row.center_name)

            if self.backend == "sqlite":
                self.cur.execute(
                    """
                    INSERT INTO monthly_data(center_id, month, year, revenue, target)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(center_id, month, year)
                    DO UPDATE SET revenue=excluded.revenue, target=excluded.target
                    """,
                    (center_id, row.month, row.year, row.revenue, row.target),
                )
            else:
                self.cur.execute(
                    """
                    INSERT INTO monthly_data(center_id, month, year, revenue, target)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT(center_id, month, year)
                    DO UPDATE SET revenue=excluded.revenue, target=excluded.target
                    """,
                    (center_id, row.month, row.year, row.revenue, row.target),
                )
            synced += 1
        return synced

    def upsert_coach_rows(self, rows: List[CoachSalaryRow]) -> int:
        synced = 0
        for row in rows:
            center_id = self._fetch_center_id(row.center_name)
            coach_id = self._fetch_coach_id(
                center_id=center_id,
                coach_name=row.coach_name,
                end_month=row.end_month,
                end_year=row.end_year,
            )

            if self.backend == "sqlite":
                self.cur.execute(
                    """
                    INSERT INTO coach_salaries(coach_id, month, year, salary)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(coach_id, month, year)
                    DO UPDATE SET salary=excluded.salary
                    """,
                    (coach_id, row.month, row.year, row.salary),
                )
            else:
                self.cur.execute(
                    """
                    INSERT INTO coach_salaries(coach_id, month, year, salary)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT(coach_id, month, year)
                    DO UPDATE SET salary=excluded.salary
                    """,
                    (coach_id, row.month, row.year, row.salary),
                )
            synced += 1
        return synced


def sync_excel_to_database() -> SyncReport:
    """Run one full sync cycle from Excel workbook to DB."""
    workbook_path = os.environ.get("EXCEL_SYNC_FILE", "instance/dashboard_sync.xlsx").strip()
    monthly_worksheet = os.environ.get("EXCEL_MONTHLY_WORKSHEET", "MonthlyData").strip()
    coach_worksheet = os.environ.get("EXCEL_COACH_WORKSHEET", "CoachSalaries").strip()

    if not os.path.exists(workbook_path):
        raise RuntimeError(f"Excel sync file not found: {workbook_path}")

    report = SyncReport(warnings=[])
    read_path, temp_copy_path = _prepare_workbook_for_sync(workbook_path)

    try:
        monthly_records = read_worksheet_as_records(workbook_path=read_path, worksheet_title=monthly_worksheet)
        report.monthly_rows_seen = len(monthly_records)
        monthly_rows, monthly_warnings = parse_monthly_data_rows(monthly_records)
        report.warnings.extend(monthly_warnings)
        report.skipped_rows += max(0, report.monthly_rows_seen - len(monthly_rows))

        coach_rows: List[CoachSalaryRow] = []
        try:
            coach_records = read_worksheet_as_records(workbook_path=read_path, worksheet_title=coach_worksheet)
            report.coach_rows_seen = len(coach_records)
            coach_rows, coach_warnings = parse_coach_salary_rows(coach_records)
            report.warnings.extend(coach_warnings)
            report.skipped_rows += max(0, report.coach_rows_seen - len(coach_rows))
        except WorksheetMissingError:
            report.warnings.append(
                f"Worksheet '{coach_worksheet}' not found; coach salary sync skipped (monthly sync still completed)."
            )
        except ValueError as exc:
            report.warnings.append(
                f"Worksheet '{coach_worksheet}' has invalid columns; coach salary sync skipped ({exc})."
            )
    finally:
        if temp_copy_path and os.path.exists(temp_copy_path):
            try:
                os.remove(temp_copy_path)
            except OSError:
                LOGGER.warning("Could not remove temporary workbook copy: %s", temp_copy_path)

    db = DatabaseClient.from_environment()
    try:
        report.monthly_rows_synced = db.upsert_monthly_rows(monthly_rows)
        if coach_rows:
            report.coach_rows_synced = db.upsert_coach_rows(coach_rows)
        db.conn.commit()
    except Exception:
        db.conn.rollback()
        raise
    finally:
        db.close()

    LOGGER.info("Excel sync complete: %s", report.as_dict())
    return report


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    report = sync_excel_to_database()
    LOGGER.info("Sync summary: %s", report.as_dict())


if __name__ == "__main__":
    main()
