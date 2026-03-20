"""Interactive Excel sync setup checker."""

from __future__ import annotations

import os
import sys

from dotenv import load_dotenv
from openpyxl import load_workbook

from excel_sync import sync_excel_to_database


load_dotenv()


def check_workbook(path: str, monthly_sheet: str, coach_sheet: str) -> bool:
    if not os.path.exists(path):
        print(f"ERROR: Excel file not found: {path}")
        return False

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = set(wb.sheetnames)
        if monthly_sheet not in names:
            print(f"ERROR: Missing worksheet '{monthly_sheet}'. Available: {', '.join(wb.sheetnames)}")
            return False
        if coach_sheet not in names:
            print(f"WARN: Worksheet '{coach_sheet}' missing. Coach salary sync will be skipped.")
        return True
    finally:
        wb.close()


def main() -> None:
    excel_file = os.environ.get("EXCEL_SYNC_FILE", "instance/dashboard_sync.xlsx").strip()
    monthly_sheet = os.environ.get("EXCEL_MONTHLY_WORKSHEET", "MonthlyData").strip()
    coach_sheet = os.environ.get("EXCEL_COACH_WORKSHEET", "CoachSalaries").strip()

    print("Excel sync setup check")
    print(f"EXCEL_SYNC_FILE={excel_file}")
    print(f"EXCEL_MONTHLY_WORKSHEET={monthly_sheet}")
    print(f"EXCEL_COACH_WORKSHEET={coach_sheet}")

    if not check_workbook(excel_file, monthly_sheet, coach_sheet):
        sys.exit(1)

    report = sync_excel_to_database()
    print("SUCCESS: Sync test completed")
    print(report.as_dict())


if __name__ == "__main__":
    main()
